from openpyxl import load_workbook

wb = load_workbook('job_listings.xlsx')
sheet = wb.active

# Data for each role (row_num: [Qualifications, Requirements, Skills])
job_data = {
    # VR / Immersive Tech
    4: {
        'qual': "BS/MS/PhD in Computer Science, Engineering, or related field; GPA 3.0+ preferred; Currently enrolled in bachelor's degree program; For Ignite: 1st/2nd year undergrad (class of 2028/2029)",
        'req': "Strong academic background in CS/engineering; Coursework in computing/engineering; 12-week minimum internship commitment; Creativity and problem-solving mentality; Willingness to develop new skills",
        'skills': "C++, Python, CUDA; OpenGL, GLES, Direct3D, Vulkan; Git, Linux; Machine learning/graphics coursework; Driver development experience; Embedded systems; Initiative, teamwork, effective communication"
    },
    5: {
        'qual': "BS/MS in Computer Science, Data Science, Electrical/Computer Engineering, or related technical field; Graduate between December 2026 and July 2027; Prior internship experience strongly preferred",
        'req': "Strong foundation in CS principles; 12-24 week internship duration (May-August 2026); Familiarity with development tools; Pass coding test (data structures, algorithms, problem-solving) and behavioral round",
        'skills': "C++, Python, Java; Unity, Unreal Engine; AR/VR development platforms; Git version control; 3D graphics; Computer vision; Strong problem-solving, adaptability, effective communication"
    },
    6: {
        'qual': "Active students or recent grads within 1 year of graduation; Experience in tech or gaming industries; Legal authorization to work in posting location; Programming portfolio demonstrating projects",
        'req': "Strong background in CS or related field; Available full-time during internship; Show curiosity and passion for interactive entertainment and 3D engine technology; Personal projects and portfolio highly valued",
        'skills': "C++ (required for engine work); Python; Unreal Engine development; Data structures & algorithms; Mathematics and physics background; Experience with scripting, exporters, data processing, packaging, deployment"
    },
    7: {
        'qual': "Bachelor's degree or higher in Computer Science, Software Engineering, Electrical Engineering, Informatics, or Design Technology; 5+ years experience for senior roles; Significant Unity engine programming experience",
        'req': "Core proficiency in programming languages; Experience with 3D modeling, animation, and visual effects; Work closely with content team; Push limits of game engine to maximize VR quality; Rapid prototyping abilities",
        'skills': "C#, C++, Java, Python; Unity3D (primary), Unreal Engine 4; VR technologies (HTC Vive Pro SDKs); 3D graphics and physics; Maintainable code architecture; VR/AR/MR, robotics, motion capture experience valued"
    },
    8: {
        'qual': "Bachelor's Degree in Computer Science or equivalent job experience; 2+ years technical engineering experience; Active US Security Clearance, or ability/willingness to obtain one",
        'req': "Architect, develop, debug, document, and deploy performant Mixed Reality applications; Invent and improve novel immersive UX techniques; Collaborate with customers, operators, MR designers; Work on high-impact, mission-critical flows",
        'skills': "C, C++, C#, Java, JavaScript, Python; Agile development methodologies; Git version control; Computer vision; Convolutional neural networks; Strong problem-solving; Quick learning of new technologies"
    },
    # Machine Learning Engineer
    9: {
        'qual': "Current or recent graduate with background in Software Development, Computer Science, Computer Engineering; BS/MS level; Must obtain work authorization in country of employment; Awards in ACM/ICPC, NOI/IOI, TopCoder, Kaggle preferred",
        'req': "Proficient coding skills; Strong algorithm & data structure basis; Effective communication and teamwork; Experience in NLP, Ranking, Ads, search engine, recommender system, distributed system, or machine learning",
        'skills': "Python, C++, or similar; NLP and CV technologies; Content understanding/matching; User behavior modeling; Video understanding; Search recommendation algorithms; Personalization systems; PyTorch, DeepSpeed, Megatron"
    },
    10: {
        'qual': "BS or equivalent in Computer Science or related field; Proven experience in AI/ML and HPC workloads; For advanced roles: Master's or PhD in CS/CE or equivalent",
        'req': "Hands-on experience with HPC infrastructure; In-depth knowledge of accelerated computing (GPU, custom silicon); Storage systems (Lustre, GPFS, BeeGFS); Scheduling/orchestration (Slurm, Kubernetes, LSF); High-speed networking (InfiniBand, RoCE)",
        'skills': "C++, Python; PyTorch (DDP, FSDP), NeMo, JAX; Docker, Enroot containers; Large-scale distributed training; GPU clusters; Kubernetes; Excellent communication and collaboration skills"
    },
    11: {
        'qual': "Graduate between Fall 2025 and Summer 2026 with degree in Computer Science, Engineering, or related subject",
        'req': "Implementation skills with general purpose programming language; Knowledge of algorithms, data structures, OOP principles; Focus on distributed data systems, optimize query execution, enhance data security/storage",
        'skills': "Python, Java, or C++; Algorithms and data structures; OOP principles; Distributed systems; Query execution optimization; Full software development lifecycle; Scalable and reliable solutions"
    },
    12: {
        'qual': "Pursuing degree in relevant field; Passion for open-source and making complex technology accessible; Interest in ML ecosystems; Background in web development and machine learning",
        'req': "Expanding Hugging Face ecosystem to web developers; Converting and optimizing models for in-browser inference; Enabling models to run in-browser at near-native speeds; Building demo applications",
        'skills': "Python, JavaScript, TypeScript; ONNX; WebGPU, WASM, WebNN; Model quantization; transformers.js and huggingface.js; In-browser inference; Low latency applications; GitHub; Open source contribution"
    },
    13: {
        'qual': "Currently pursuing bachelor's or completed within 1 year in Computer Science or related field; For PhD track: pursuing PhD or Master's degree; BS/MS required; Must commit to onboarding by end of 2026",
        'req': "Solid knowledge in machine learning, pattern recognition, NLP, data mining, or computer vision; Firm understanding of data structures and algorithms; Great communication and teamwork; Develop ML algorithms for business risk management",
        'skills': "C++, Go, Python; CUDA; PyTorch, DeepSpeed, Megatron, vllm; Distributed systems; Large-scale data processing; Parallel computing; LLM and generative AI experience; Deep learning frameworks"
    },
    # Analytics / Data Engineer
    14: {
        'qual': "Currently has or obtaining Bachelor's degree in Computer Science, Computer Engineering, or relevant technical field; Degree completed prior to joining Meta; 2026 graduates eligible; Prior internship experience strongly preferred",
        'req': "Architect, implement, deploy new data models and processes in production; Perform data analysis for business insights; Interface with Engineers, Product Managers, Product Analysts; Fast-paced work environment; Curious, self-driven, analytical",
        'skills': "SQL (joins, aggregations, window functions, complex queries); Python; Java; Data modeling; ETL processes; Big data technologies; Data visualization tools; Scalable database schema design"
    },
    15: {
        'qual': "Pursuing degree in Computer Science, Data Engineering, or related field; Must keep camera on during video interviews; Required to attend in-person onboarding; Excited to learn and develop skills",
        'req': "Partner with engineers and analysts to design/build reliable datasets; Ensure data quality; Make data accessible across company; Build and maintain data pipelines connecting different data sources; Optimize for performance and scalability",
        'skills': "Data pipeline development; Data structures; SQL; Data modeling; ETL concepts; Collaborative skills; Problem-solving; Cross-functional teamwork"
    },
    16: {
        'qual': "Currently enrolled in M.S in Computer Science, Engineering or related STEM field; Demonstrated critical thinking skills; Ability to commit to minimum 12-week summer internship; Work authorization may be required",
        'req': "Deliver full-stack data solutions across entire data processing pipeline; Work through all stages of data solution lifecycle; Create conceptual, logical and physical data model designs; Architect ETL, reporting and analytics; Build exec-facing dashboards",
        'skills': "Go, Python, SQL, Java; Modern enterprise data architectures; Design patterns; Data tool sets; SQL/NoSQL; ETL design; Distributed systems (Spark); Reporting and analytics; Cross-functional collaboration"
    },
    17: {
        'qual': "Actively enrolled in accredited college/university; 3rd/4th year Undergraduates, Masters, or PhD; Major in Computer Science, Physics, Math, or related field",
        'req': "Required coursework: Algorithms, Data Structures, Operating Systems; Recommended: database systems, distributed systems; Experience working as part of a team; Dedication and passion for technology",
        'skills': "Python, Java, SQL; C++ or Java for coding assessments; Data warehousing concepts; Cloud platforms (AWS, Azure, GCP); Git version control; DevOps (Jenkins, Ansible, Terraform); Containerization (Kubernetes, Docker)"
    },
    18: {
        'qual': "Bachelor's or Master's degree in Computer Science or directly related field by Summer 2026, or equivalent work experience; Some programming experience through side projects or classwork",
        'req': "Experience from internships or multi-person coding projects; Ability to learn unfamiliar systems through research and mentorship; Understanding of how to safely update production systems; Code review experience",
        'skills': "Java, Ruby, JavaScript, Scala, Go; Frontend, backend, or full-stack experience; HTTP, service layers, REST APIs; Algorithms, data structures, object-oriented design; Large codebase navigation"
    },
    19: {
        'qual': "Currently pursuing relevant degree; Available for full-time 40 hours/week; Must be able to work 2 days/week onsite in Needham, MA (hybrid); No relocation or housing assistance provided",
        'req': "Gain hands-on exposure to tools, frameworks, and practices powering modern data ecosystem; Work on meaningful projects; Sharpen skills and deepen knowledge; Present contributions to SharkNinja leaders",
        'skills': "Analytical thinking; Strong problem-solving; Attention to detail; Data modeling; ETL concepts; Big data tools; Git/version control; Clear technical communication; Cross-functional collaboration"
    },
    20: {
        'qual': "Pursuing relevant degree; Stable internet connection and access to PC/laptop; Mission-driven with 'all-in' disposition; Ability to excel at independent and collaborative work",
        'req': "Effective communication in remote cross-functional environment; Fast-paced, high-growth organization with evolving priorities; Clear and candid communication; Continuous learning mindset; Remote-first work environment",
        'skills': "Data engineering fundamentals; Communication skills; Remote collaboration tools; Self-motivation; Adaptability to change; Problem-solving"
    },
    # AI Engineer
    21: {
        'qual': "Bachelor's or Master's degree in Computer Science, Computer Engineering, relevant technical field, or equivalent practical experience; 0-1 years of professional experience in software engineering or relevant field",
        'req': "Own development of customer-facing ChatGPT and OpenAI API features end-to-end; Talk to users to understand problems and design solutions; Collaborate with cross-functional team; Move fast in loosely defined environment",
        'skills': "JavaScript, React; Python (backend); Relational databases (Postgres/MySQL); Interest in AI/ML (direct experience not required); Self-starter mindset; User experience focus; Handle competing priorities"
    },
    22: {
        'qual': "Graduation date Fall 2025 or Spring 2026; Bachelor's degree (or equivalent) in Computer Science, EECS, Computer Engineering, or Statistics",
        'req': "Product engineering experience building web apps full-stack; Integrating with APIs and services; Talking to customers; Previous Product/Software Engineering Internship; Track record of shipping high-quality products at scale; Experience with large data volumes",
        'skills': "TypeScript, React, Python, MongoDB; Full-stack web development; API integration; Building tools for labeling platforms; Fraud-detection systems; Matching algorithms; UI/UX tooling; ML algorithms"
    },
    23: {
        'qual': "Bachelor's degree or higher in Computer Science or related field (or equivalent experience)",
        'req': "Practical experience in AI, machine learning, or agent frameworks; Exposure to microservices, web apps, databases; Containerization (Docker), Kubernetes, CI/CD pipelines; Distributed messaging systems (Kafka) or event-driven architectures; Proactive approach",
        'skills': "Python (required), JavaScript; Software engineering principles; OOP/functional programming; High-performance maintainable code; LangChain, OpenAI Functions; SQL/NoSQL databases; Docker; Kubernetes; Kafka; CI/CD"
    },
    24: {
        'qual': "Currently pursuing PhD in Computer Science or related field; Or BS/MS in CS, Linguistics, Statistics, Biostatistics, Applied Mathematics, Operations Research, Economics, or Natural Sciences; Must intend to resume degree after research role",
        'req': "Experience as researcher (internships, full-time, or lab); Publishing papers in major conferences/journals; Experience contributing to research communities; Strong academic record; Proven initiative in research settings",
        'skills': "Python, Java, C++, MATLAB; NLU, HCI, Generative Media, Computer Vision, ML, Deep Learning, Optimization, Quantum Information Science, Data Science; Research publication experience; Effective English communication"
    },
    # Data Scientist (Consumer)
    25: {
        'qual': "Junior: 0-2 years experience; Degree in computer science, math, economics, or statistics; Located in Central European or Eastern Standard/Daylight time zones",
        'req': "Support product decision making, goal setting, metric development, and A/B testing; Use analytical and statistical approaches to derive insights; Partner with cross-functional teams; Communicate complex findings to diverse stakeholders",
        'skills': "SQL and Python/R fluency; Statistical packages; Large complex data sets; Data preparation and exploratory analysis; Data visualization (Tableau, Looker Studio); A/B testing; Personalization algorithms"
    },
    26: {
        'qual': "MS or PhD in Statistics, Econometrics, Computer Science, Physics, or related field preferred; 5+ years specific experience; Strong analytical and communication skills",
        'req': "End-to-end analytics: designing and analyzing randomized experiments, building causal-inference pipelines, informing high-stakes studio and product decisions; Guide content greenlighting, personalization, and user growth initiatives",
        'skills': "SQL, Python, R, Java, or Scala; PySpark or Scala Spark; Distributed analytics (Spark, Presto, Hive); A/B testing; Tableau; Machine learning models; Causal inference; Data visualization (D3, R Shiny)"
    },
    27: {
        'qual': "Bachelor's in quantitative field; Master's or PhD is a plus; For staff roles: 6-9+ years experience; Advanced degree in Operations Research, Economics, Mathematics, Physics preferred",
        'req': "Analyze large datasets to uncover insights; Develop predictive models; Design experiments (A/B tests); Collaborate with product and engineering teams; Enhance user experience, pricing, search algorithms, fraud detection",
        'skills': "Python or R; SQL; Machine learning (K-means, KNN, linear regression, SVM, decision trees, random forests); Statistical analysis; Data visualization; Exploratory data analysis; Cross-functional collaboration"
    },
    28: {
        'qual': "Bachelor's degree in Computer Science, Mathematics, Statistics, or related technical field; For senior: Ph.D., M.S. or Bachelor's in quantitative fields; Expected graduation Fall 2026 or Spring 2027 for internships",
        'req': "Design, implement, and track core metrics to analyze product performance; Translate ambiguous problems into data-driven analyses; Experimental design and analysis (A/B and market-level experiments); Strong business and product sense",
        'skills': "SQL, Python, or R; Tableau, Mixpanel, Looker; Hadoop, Spark; Statistical/ML methods (regression, classification, causal inference, time series); A/B testing; Strong storytelling abilities"
    },
    29: {
        'qual': "BA/BS in Computer Science, Applied Math, Physics, Engineering, Statistics, Economics or equivalent; MS/MA preferred; For PhD roles: PhD in Data Science, Biostatistics, Mathematics or related; 4+ years experience typically required",
        'req': "Develop modeling solutions to identify malicious user behavior; Build dashboards to understand root causes of metric changes; Evaluate A/B experiments for product feature launches; Execute complex cross-functional projects",
        'skills': "SQL (joins, aggregations, window functions); Python or R; Big data (Spark, Hive, Airflow); A/B testing; Causal inference; Machine learning and Neural Networks; Econometric techniques; Statistical analysis"
    },
    30: {
        'qual': "Bachelor's/Master's degree in Computer Science or equivalent; 3+ years industry experience (6+ for lead roles); 7+ years for staff positions in fast-paced, data-driven environment",
        'req': "Develop deep understanding of engagement ecosystem and key product surfaces; Deep dives on engagement metrics impact on monetization and retention; Design ML and evaluation frameworks; Mentor junior scientists",
        'skills': "Python/R; SQL/Spark; Recommendation systems; Causal inference; Deep learning (user modeling, image recognition, NLP); A/B testing; Data visualization (Tableau, Looker); Quantitative programming"
    },
    # Data Scientist (ML)
    31: {
        'qual': "PhD in Computer Science, Linguistics, Statistics, Applied Mathematics, or related field; Or MS with significant research experience; Strong academic record; Currently enrolled or recently completed",
        'req': "Experience as researcher in internships, full-time, or lab settings; Publishing papers in major conferences/journals; Contributing to research communities; Proven initiative in research settings",
        'skills': "Python, Java, C++, MATLAB; NLU, Computer Vision, ML, Deep Learning, Optimization, Quantum Information Science; Research publication experience; Statistical modeling; Neural network architectures"
    },
    32: {
        'qual': "Bachelor's minimum; Master's preferred; PhD in AI, ML, computer science, or related technical fields ideal; First author publications at NeurIPS, CVPR, ICML, ICLR, ICCV, ACL valued",
        'req': "Proven track record of publication; Direct experience in generative AI and LLM research; SWE coding interviews (LeetCode medium); System design for scalable ML systems; Role-related knowledge interviews",
        'skills': "PyTorch, TensorFlow, Scikit-learn, Pandas; C/C++ or Python; Linear algebra, calculus, statistics, probability; CNNs, RNNs, transformers; GPU architecture; Jupyter, Matplotlib; Scientific computing"
    },
    33: {
        'qual': "PhD or MS in Computer Science, Statistics, Electrical Engineering, or related field; Strong publication record; Experience with graphics and simulation research",
        'req': "Research and develop AI algorithms for graphics and gaming; Work on simulation environments; Accelerate robotics development; Collaborate with simulation engineers, researchers, and artists",
        'skills': "C++, Python, CUDA; Deep learning frameworks (PyTorch, TensorFlow); Computer graphics; 3D simulation; Generative models; Neural rendering; Physics-based simulation; Research publication experience"
    },
    34: {
        'qual': "PhD or MS in Computer Science, Statistics, Electrical Engineering, Applied Math, Operations Research, Econometrics or equivalent; Currently enrolled pursuing PhD in ML, AI, CS, statistics, or scene semantic understanding",
        'req': "Research and develop attribution, media mix modeling, budget optimization, personalization, causal analysis, time series analysis; Use statistical/econometric methods, predictive models, experimental design, optimization",
        'skills': "Python, Scala, Java, C; PyTorch, TensorFlow; R, MATLAB, scikit-learn; SQL; Statistical modeling; Machine learning; Deep learning; Computer vision; Audio processing; Generative models; Strong communication"
    },
    35: {
        'qual': "BS degree in Computer Science, Statistics, Economy or other relevant majors with 0-1 year of work experience; Commit to onboarding by end of year 2026",
        'req': "Develop scientific metrics to measure E-Commerce governance success; Develop credit system for merchants and creators; Work with engineers and product managers; Drive project execution and iteration",
        'skills': "SQL, R, Python; Good self-driving ability; Team cooperation; Strong sense of responsibility; Willingness to learn new knowledge; Data analysis; Metric development"
    },
    36: {
        'qual': "Ph.D. in computer science, engineering, or relevant degree; 5-7+ years experience in Machine Learning and NLP/NLG/Speech; Publications in top-tier ML conferences; Proven track record in applied ML research",
        'req': "Train large scale language and multimodal models; Deploy compact neural architectures on device; Learn policies personalized to user in privacy preserving manner; Data preparation, prototyping, production implementation",
        'skills': "Python, Swift, C++; PyTorch, TensorFlow, Core ML; LLMs, VLMs, diffusion models; NLP, NLG, Speech; Privacy-preserving ML (differential privacy, secure multiparty compute); Generative models; Cross-functional communication"
    },
    # Analytics Engineer
    37: {
        'qual': "SQL proficiency required; At least 6 months experience working with dbt (Core or Cloud); Familiarity with branching strategies, basic git commands, and pull requests",
        'req': "Build, test, and maintain models for data accessibility; Apply engineering principles to analytics infrastructure; Set up/audit/refactor client dbt projects; Communicate findings to broad range of stakeholders",
        'skills': "SQL (advanced); dbt (Core and Cloud); Git/version control; Data warehouse fundamentals (Snowflake, BigQuery); Data modeling; ETL concepts; Performance tuning; Documentation"
    },
    38: {
        'qual': "Degree in computer science, math, economics, or statistics preferred; Experience with data modeling and analytics engineering; Located in collaboration-friendly timezone",
        'req': "Design and maintain analytics-friendly datasets; Support experimentation and product analytics; Partner with cross-functional teammates; Optimize pipelines for performance and scalability",
        'skills': "SQL; Python; dbt; Data modeling; ETL pipelines; Snowflake or BigQuery; Git version control; Data visualization; A/B testing support; Cross-functional collaboration"
    },
    39: {
        'qual': "Bachelor's degree minimum; Master's preferred; 2-5+ years experience in analytics engineering or data engineering; Strong SQL and data modeling skills",
        'req': "Build and maintain data pipelines; Create analytics-ready datasets; Support finance and corporate analytics; Work with Snowflake and modern data stack; Ensure data quality and documentation",
        'skills': "SQL (advanced); dbt; Snowflake; Python; Spark; Data modeling; Dimensional modeling; ETL design; Data visualization (Tableau); Git version control"
    },
    40: {
        'qual': "5+ years experience in Analytics Engineering, Data Science, Data Engineering, or similar; Master's or PhD is a plus; Degree in quantitative field (Statistics, Econometrics, CS, Engineering, Mathematics, Data Science)",
        'req': "Sit at intersection of data science, product analytics, and data engineering; Architect, build, and launch data models and pipelines; Design and implement metrics and dimensions; Interface with Analytics Engineers, Data Scientists, Business Partners",
        'skills': "SQL (advanced); Spark, Presto, Hive; Python (Scala preferred); Data modeling; Dimensional modeling; Tableau, Superset; Metric development; Schema design; Geospatial data experience valued"
    },
    41: {
        'qual': "2+ years relevant industry experience; Advanced degree (M.S. or Ph.D.) in CS, Software Engineering, AI, ML, Statistics, Economics, Physics preferred; Bachelor's required",
        'req': "Build data models and pipelines; Develop metrics and dashboards; Support data-driven decision-making; Translate complex data insights into business strategies; Work in fast-paced, dynamic environment",
        'skills': "SQL; Python; Looker; Data modeling; ETL processes; MySQL, Snowflake, Redshift; Tableau; A/B testing; Experimental design; Cross-functional collaboration"
    },
    42: {
        'qual': "5+ years hands-on experience in BI, Data Science, or Data Engineering; Track record of success in fast-paced environment; Proven accomplishments using needed skills",
        'req': "Build on Instacart's Data Foundations; Design, create, and maintain analytics-friendly datasets; Collaborate with stakeholders across company; Develop data metrics, validation, documentation, and tooling",
        'skills': "SQL; dbt; Data modeling; Python; Spark; Data pipelines; Cross-functional collaboration; Self-driven strategic thinking; Technical documentation"
    },
    # Data Engineer
    43: {
        'qual': "Bachelor's degree in Computer Science, Engineering, or related field; Experience with distributed systems and cloud platforms; Graduate or 1-2 years experience",
        'req': "Build and maintain data platform infrastructure; Design scalable data pipelines; Work with Spark and distributed processing; Ensure data quality and reliability; Collaborate with data scientists and analysts",
        'skills': "Python, Java, Scala; SQL; Apache Spark; Delta Lake; Databricks platform; Cloud platforms (AWS, Azure, GCP); ETL development; Data modeling; Git version control"
    },
    44: {
        'qual': "Bachelor's degree in Computer Science, Engineering, or related field; Understanding of data warehousing and cloud platforms; 1-3 years experience or new grad with strong fundamentals",
        'req': "Build and optimize data pipelines; Work with Snowflake data warehouse; Design efficient data models; Ensure data quality and governance; Collaborate with analytics and BI teams",
        'skills': "SQL (advanced); Python; Snowflake; Cloud platforms (AWS, Azure, GCP); ETL development; Data modeling; dbt; Airflow; Git version control; Data warehouse optimization"
    },
    45: {
        'qual': "Bachelor's degree in Computer Science or related field; Understanding of ETL and data modeling; Proficient in scripting/programming; New grad or 1-2 years experience",
        'req': "Build data warehousing layers; Lead design and delivery of large-scale data solutions; Build scalable data infrastructure; Understand distributed systems; Handle large volume and velocity data",
        'skills': "SQL; Python; Java; AWS services (S3, Redshift, Glue, Lambda, Kinesis, DynamoDB); ETL development; Data modeling; Spark; Data pipelines; Distributed systems"
    },
    46: {
        'qual': "Bachelor's degree in Computer Science, Mathematics, or related field; 1+ years experience with Google Cloud recommended; Basic SQL proficiency; Understanding of data processing",
        'req': "Design, build, deploy data processing systems; Work with Google Cloud data services; Create and manage data pipelines; Ensure data quality and security; Collaborate with analysts and data scientists",
        'skills': "Python, Java; SQL; Google Cloud (BigQuery, Dataflow, Data Fusion, Dataproc, Pub/Sub); Apache Beam; Data modeling; ETL development; Machine learning basics; Data visualization"
    },
    47: {
        'qual': "Bachelor's degree in Computer Science, Mathematics, Statistics, or related field; 2+ years experience preferred; Strong SQL and Python skills",
        'req': "Build robust data pipelines and infrastructure; Design data models; Ensure data integrity and quality; Collaborate with data scientists and analysts; Own core data pipelines powering top-line metrics",
        'skills': "SQL; Python, Ruby, Bash; MySQL, PostgreSQL; Hadoop ecosystem (MapReduce, HDFS, Hive, Spark, Presto); Airflow, Oozie; Kafka; Data modeling; Performance tuning"
    },
    48: {
        'qual': "Bachelor's degree in Computer Science or related field; 2+ years experience in data engineering; Strong problem-solving abilities; Passion for working with data",
        'req': "Build data transport, collection, and storage systems; Own core data pipelines; Evolve data models; Architect and build scalable data pipelines; Work with Analytics, Data Science, Marketplace teams",
        'skills': "SQL; Python, Ruby, Bash; MySQL, PostgreSQL; Hadoop ecosystem (Spark, Hive, Presto, Parquet); Airflow, Oozie, Azkaban; Data modeling; Large-scale data processing"
    },
    49: {
        'qual': "Bachelor's or Master's degree in Computer Science or related field by Summer 2026; Some programming experience; Experience from internships or coding projects",
        'req': "Work on cross-functional projects; Give meaningful code review feedback; Ensure systems scale to meet user needs; Build skills to own projects end-to-end; Learn technical leadership",
        'skills': "Java, Ruby, JavaScript, Scala, Go; SQL; Data modeling; ETL development; Distributed systems; Payment systems knowledge; Code review practices; Large codebase navigation"
    }
}

# Update each row
for row_num, data in job_data.items():
    sheet.cell(row=row_num, column=7, value=data['qual'])
    sheet.cell(row=row_num, column=8, value=data['req'])
    sheet.cell(row=row_num, column=9, value=data['skills'])

wb.save('job_listings.xlsx')
print(f'Updated {len(job_data)} rows with Qualifications, Requirements, and Skills')
